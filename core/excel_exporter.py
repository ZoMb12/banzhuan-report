import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Microsoft YaHei", size=10)
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def export_to_excel(target_windows, conversion_rate: float, output_dir: str = None):
    """Export target window results to Excel — one row per time-point."""
    if output_dir is None:
        output_dir = config.EXPORT_DIR
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"搬砖报表_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "目标饰品明细"

    # ---- Title row ----
    ws.merge_cells("A1:J1")
    ws["A1"] = f"搬砖报表 — 目标饰品 — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ---- Config row ----
    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"考察天数: {config.DEFAULT_STABLE_DAYS}天 | "
        f"波动阈值: {config.DEFAULT_VOLATILITY_THRESHOLD * 100:.0f}% | "
        f"汇率: 1 USD = {conversion_rate} CNY | "
        f"仅含均价差>0的目标饰品"
    )
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # ---- Headers (row 3) ----
    headers = [
        "序号", "饰品名称", "窗口起始", "窗口结束",
        "BUFF日期", "BUFF价格(¥)",
        "Steam日期", "Steam价格($)", "Steam价格(¥)",
        "差价(¥)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 25

    # ---- Data: group by (item, window), then list each time-point ----
    row = 4
    seq = 1
    # Sort: by item name then window start
    sorted_windows = sorted(target_windows, key=lambda w: (w.item_name, w.window_start))

    for wr in sorted_windows:
        # Section header row for this item-window
        ws.merge_cells(f"A{row}:J{row}")
        window_info = (
            f"{wr.item_name}  |  窗口: {wr.window_start} ~ {wr.window_end}  |  "
            f"BUFF均价: ¥{wr.buff_avg_price:.2f}  |  "
            f"Steam均价: ${wr.steam_avg_price_usd:.2f}  |  "
            f"均价差: ¥{wr.avg_diff:+.2f}  |  "
            f"利润率: {wr.avg_profit_rate * 100:.2f}%"
        )
        cell = ws.cell(row=row, column=1, value=window_info)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True)
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        # Time-point rows
        if wr.date_pairs:
            for dp in wr.date_pairs:
                values = [
                    seq,
                    wr.item_name,
                    wr.window_start.strftime("%Y-%m-%d") if isinstance(wr.window_start, date) else str(wr.window_start),
                    wr.window_end.strftime("%Y-%m-%d") if isinstance(wr.window_end, date) else str(wr.window_end),
                    dp["buff_date"].strftime("%Y-%m-%d") if isinstance(dp["buff_date"], date) else str(dp["buff_date"]),
                    round(dp["buff_price"], 2),
                    dp["steam_date"].strftime("%Y-%m-%d") if isinstance(dp.get("steam_date"), date) else (str(dp["steam_date"]) if dp.get("steam_date") else "N/A"),
                    round(dp["steam_price_usd"], 2) if dp.get("steam_price_usd") else "N/A",
                    round(dp["steam_price_cny"], 2) if dp.get("steam_price_cny") else "N/A",
                    round(dp["diff"], 2) if dp.get("diff") is not None else "N/A",
                ]

                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                ws.row_dimensions[row].height = 20
                seq += 1
                row += 1
        else:
            cell = ws.cell(row=row, column=1, value="无时间点明细数据")
            ws.merge_cells(f"A{row}:J{row}")
            cell.font = Font(name="Microsoft YaHei", size=10, color="999999")
            cell.alignment = Alignment(horizontal="center")
            row += 1

        row += 1  # blank row between item-windows

    # ---- Column widths ----
    col_widths = [6, 35, 13, 13, 13, 14, 13, 14, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ---- Freeze panes ----
    ws.freeze_panes = "A4"

    # ---- Summary sheet ----
    ws2 = wb.create_sheet("汇总统计")
    unique_items = len(set(w.item_id for w in target_windows))
    total_pairs = sum(len(w.date_pairs) for w in target_windows)

    summary_data = [
        ["指标", "数值"],
        ["目标饰品数", unique_items],
        ["目标窗口数", len(target_windows)],
        ["总时间点数", total_pairs],
        ["考察天数", f"{config.DEFAULT_STABLE_DAYS}天"],
        ["波动阈值", f"{config.DEFAULT_VOLATILITY_THRESHOLD * 100:.0f}%"],
        ["汇率", f"1 USD = {conversion_rate} CNY"],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = HEADER_FONT
            cell_a.fill = HEADER_FILL
            cell_b.font = HEADER_FONT
            cell_b.fill = HEADER_FILL
        else:
            cell_a.font = DATA_FONT
            cell_b.font = DATA_FONT
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        cell_b.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 25

    wb.save(filepath)
    return filepath
